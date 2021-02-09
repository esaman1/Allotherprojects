import openpyxl
import os

# os.chdir(r'C:\Users\ahame\OneDrive\Documents')
# workbook = openpyxl.load_workbook("Book1.xlsx")

# sheet = workbook.get_sheet_by_name("Sheet1")
# cell = sheet["A1"]
# print(sheet.cell(1,1).value)
# print(cell.value)



wb = openpyxl.Workbook()
# print(wb.get_sheet_names())

sheet = wb.get_sheet_by_name("Sheet")
sheet['A1'] = 42
sheet['A2'] = 'Hello'
sheet2 = wb.create_sheet()
sheet2.title = "Example Sheet"
print(wb.get_sheet_names())

os.chdir(r'C:\Users\ahame\OneDrive\Documents')
wb.save('example.xlsx')
